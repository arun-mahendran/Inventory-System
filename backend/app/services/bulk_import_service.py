from sqlalchemy.orm import Session

from openpyxl import load_workbook

from app.schemas.parcel import ParcelCreate

from app.services.parcel_service import (
    create_parcel
)


def import_parcels_from_excel(
    db: Session,
    file
):

    workbook = load_workbook(file.file)

    sheet = workbook.active

    imported = 0

    skipped = 0

    errors = []

    for row_number, row in enumerate(
        sheet.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        try:

            parcel = ParcelCreate(

                tracking_number=row[0],

                customer_name=row[1],

                phone=str(row[2]),

                email=row[3],

                address=row[4],

                pincode=str(row[5]),

                amount=float(row[6]),

                payment_method=row[7],

                payment_status=(
                    "Paid"
                    if row[7] == "Prepaid"
                    else "Pending"
                )
            )

            create_parcel(
                db,
                parcel
            )

            imported += 1

        except Exception as e:

            skipped += 1

            errors.append(
                f"Row {row_number}: {str(e)} "
                f"(Pincode: {row[5]})"
        )

    return {

        "total_rows":
            imported + skipped,

        "imported":
            imported,

        "skipped":
            skipped,

        "errors":
            errors
    }